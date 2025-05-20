# -*- coding: utf-8 -*-
"""
Created on Wed Mar 19 20:24:13 2025

@author: Levi
"""

from typing import List, Dict
from datetime import datetime
import numpy as np
import openpyxl

def open_worksheet(file_path):
    """
    Returns the first worksheet of the excel file at the given filepath.

    Args:
        file_path (str): The filepath to the excel file.

    Returns:
        (openpyxl.worksheet): the first worksheet of the given excel file.
    """
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(filename=file_path)
    ws = workbook.active
    return ws

def extract_availability(ws) -> np.ndarray:
    """
    Extracts the availability information from a given worksheet and returns it
    as a numpy array.

    Availability is represented as:
    - -1 for None (available)
    - 0 for "X" or "b" (unavailable)
    - 1 for "ü" (present)

    Args:
        ws (openpyxl.worksheet): The worksheet containing the schedule.

    Returns:
        np.ndarray: A numpy array where each row represents a person and each
                    column represents a time slot or availability, with values
                    indicating availability.
    """
    # Get the number of rows and columns in the availability section
    n_rows = len(list(ws.rows)) - 3 - 8  # Skip header, counting and legend rows
    n_cols = len(list(ws.columns)) - 3  # Skip person, counting and preference columns

    # Initialize the availability array with zeros (assuming available by default)
    availability = -np.ones((n_rows // 2, n_cols), dtype=int)

    # Iterate over rows to populate the availability array
    for i, row in enumerate(ws.values):
        # Skip rows with dates, services, tasks, or counts (rows that are not
        # availability data)
        if i < 2 or i % 2 == 1 or i > n_rows + 1:
            continue

        # Iterate over columns for each person (skipping the first and last 2 columns)
        for j, value in enumerate(row[1:-2]):
            if value is None:
                availability[(i - 2) // 2][j] = 0  # Available
            elif value == "x":
                availability[(i - 2) // 2][j] = -1  # Unavailable
            elif value == "b":
                availability[(i - 2) // 2][j] = -1  # Unavailable
            elif value == "ü":
                availability[(i - 2) // 2][j] = 1  # Present
            else:
                # Log invalid values and ignore them
                print(f"Invalid value '{value}' at position ({i}, {j}). It will "
                      "be ignored.")
                availability[(i - 2) // 2][j] = 0  # Set as available

    return availability

def extract_services(ws) -> Dict[datetime, List[float]]:
    """
    Extracts service data from the worksheet and returns a dictionary mapping
    dates to services.

    The services are represented by:
    - 0 for service 'o'
    - 1 for service 'a'
    - 0.5 for other service types.

    Args:
        ws (openpyxl.worksheet): The worksheet containing the service information.

    Returns:
        dict: A dictionary where the keys are dates and the values are lists of
              service types.
    """
    # Get the number of columns containing service data
    n_cols = len(list(ws.columns)) - 3

    # Initialize the services dictionary
    services_dict = {}

    # Iterate over the first two rows (date and service type) in each column
    for col in ws.iter_cols(min_col=2, max_col=1 + n_cols, max_row=2,
                             values_only=True):
        # Determine service type based on the value in the second row
        if col[1] == 'o':
            service = 0
        elif col[1] == 'a':
            service = 1
        else:
            service = 0.5  # Default to 0.5 for other service types

        # Add the service to the dictionary, with date as key and service as value
        if col[0] is not None:
            date = col[0]
            services_dict[date] = [service]
        else:
            services_dict[date].append(service)

    return services_dict

def extract_prefs(ws) -> List[str]:
    """
    Extracts preference data from the worksheet and returns it as a list of
    preferences.

    Preferences are assumed to be present in the last column, alternating
    between even-indexed rows containing preference values.

    Args:
        ws (openpyxl.worksheet): The worksheet containing preference data.

    Returns:
        List[str]: A list of preferences for each person.
    """
    prefs = []

    # Get the last column excluding the header and bottom line
    col = list(*ws.iter_cols(min_col=len(list(ws.columns)), min_row=3,
                             max_row=len(list(ws.rows)) - 1, values_only=True))

    for i, value in enumerate(col):
        if i % 2 == 0:  # Only collect values in even-indexed rows (preferences)
            prefs.append(value)

    return prefs

if __name__ == "__main__":
    # Define the file path of the Excel sheet
    FILENAME = "test_kerkenraadsrooster.xlsx"
    worksheet = open_worksheet(FILENAME)

    print(extract_availability(worksheet))
    print(extract_services(worksheet))
    print(extract_prefs(worksheet))
