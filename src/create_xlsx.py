# -*- coding: utf-8 -*-
"""
Created on Sat Jan  4 10:42:23 2025

@author: Levi
"""

import os

from datetime import datetime
from typing import Optional, Dict, List, Literal
from pydantic import BaseModel, EmailStr

import numpy as np

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.cell import get_column_letter  # pylint: disable=no-name-in-module
from openpyxl.worksheet.worksheet import Worksheet

from date_utils import get_church_dates, get_services, services_to_list


class Person(BaseModel):
    """
    This class represents a person in the schedule.
    """
    ambt: str
    email: Optional[EmailStr] = None
    telnr: Optional[str] = None
    voorkeur: Literal["geen", "om de week", "ochtend", "avond"]

Persons = Dict[str, Person]
Services = Dict[datetime, List[int]]

# Initialize kerkenraad
kerkenraad = {
    "C. Huisman": {
        "ambt": "ouderling",
        "email": "huisman@ziggo.nl",
        "telnr": "06-25055447",
        "voorkeur": "om de week",
        },
    "K. Dorresteijn": {
        "ambt": "ouderling",
        "email": "keesdorresteijn@outlook.com",
        "telnr": "06-22555003",
        "voorkeur": "geen",
        },
    "G.J. Verburg": {
        "ambt": "ouderling",
        "email": "gjverburg@outlook.com",
        "telnr": "06-10059937",
        "voorkeur": "ochtend",
        },
    "P. van den Broek": {
        "ambt": "ouderling",
        "email": "vdbroek@live.nl",
        "telnr": "06-19456958",
        "voorkeur": "geen",
        },
    "J.W. Hofland": {
        "ambt": "ouderling",
        "email": "jw.hofland@ziggo.nl",
        "telnr": "06-10271823",
        "voorkeur": "avond",
        },
    "R.G. van Lieshout": {
        "ambt": "ouderling",
        "email": "ouderlingwijk4@hervormdlopik.nl",
        "telnr": "06-55725327",
        "voorkeur": "geen",
        },
    "D. de Jong": {
        "ambt": "kerkrentmeester",
        "email": "info@bouwbedrijfdejonglopik.nl",
        "telnr": "06-51052856",
        "voorkeur": "avond",
        },
    "N. Schep": {
        "ambt": "kerkrentmeester",
        "email": "n.schep@outlook.com",
        "telnr": "06-20392712",
        "voorkeur": "ochtend",
        },
    "A. Buitenhuis": {
        "ambt": "kerkrentmeester",
        "email": "aronbuitenhuis@gmail.com",
        "telnr": "06-11820506",
        "voorkeur": "ochtend",
        },
    "H. de Groot": {
        "ambt": "diaken",
        "email": "de.groot.hans@hotmail.com",
        "telnr": "06-12102581",
        "voorkeur": "om de week",
        },
    "K. Goedhart": {
        "ambt": "diaken",
        "email": "famgoedhart@ziggo.nl",
        "telnr": "06-51082473",
        "voorkeur": "ochtend",
        },
    "M. Visser": {
        "ambt": "diaken",
        "email": "martijnvisser76@hotmail.com",
        "telnr": "06-23372557",
        "voorkeur": "geen",
        },
    "G. de Vor": {
        "ambt": "diaken",
        "email": "gerritdevor@outlook.com",
        "telnr": "06-36436022",
        "voorkeur": "ochtend",
        },
}

# Initialize sides
thin = Side(border_style="thin")
thick = Side(border_style="medium")


############################################
# ----------- Add first column ----------- #
############################################

def add_header(ws: Worksheet, row: int, col: int, name: str,
               col_width: Optional[int] = None) -> Worksheet:
    """
    Adds a formatted header to a specific cell in the worksheet.

    Args:
        ws (Worksheet): The worksheet to add the header to.
        row (int): Row index for the header.
        col (int): Column index for the header.
        name (str): Header text.
        col_width (Optional[int]): Optional width to set for the column.

    Returns:
        Worksheet: The updated worksheet.
    """
    col_letter = get_column_letter(col)
    cell = ws.cell(row, col, name)
    cell.alignment = Alignment(horizontal= "center", vertical="center")
    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ws.merge_cells(f"{col_letter}{row}:{col_letter}{row + 1}")

    if col_width is not None:
        ws.column_dimensions[col_letter].width = col_width

    return ws

def add_first_col(ws: Worksheet, persons: Persons) -> Worksheet:
    """
    Adds the first column to the worksheet with member names and phone numbers.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Add version
    add_header(ws, 1, 1, "Versie 1")

    # Add names and phone numbers
    for i, name in enumerate(persons):
        ws.cell(2*(i + 1) + 1, 1, name)
        ws.cell(2*(i + 2), 1, persons[name]["telnr"])

    # Add closing cell
    ws.cell(2*(len(persons) + 1) + 1, 1, "Aantal kerkenraadsleden:")

    return ws

def autoscale_col(ws: Worksheet, col_letter: str) -> Worksheet:
    """
    Adjusts the column width to fit its longest content.

    Args:
        ws (Worksheet): The worksheet to update.
        col_letter (str): The letter of the column to autoscale.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initialize column and its new width
    col = ws[col_letter]
    new_col_width = 0

    # Update the new column width to be the length of the longest entry
    for cell in col:
        entry_length = len(str(cell.value))
        if entry_length > new_col_width:
            new_col_width = entry_length
    new_col_width -= 2

    ws.column_dimensions[col_letter].width = new_col_width

    return ws


##################################################
# ----------- Add dates and services ------------#
##################################################

def add_dates(ws: Worksheet, services: Services) -> Worksheet:
    """
    Adds formatted dates to the worksheet header row.

    Args:
        ws (Worksheet): The worksheet to update.
        services (Services): Contains the services for each date.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initialize starting column
    column = 2

    for date, service_list in services.items():
        # Add the date to the cell
        cell = ws.cell(1, column, date.strftime("%d-%m"))

        # Merge cells if there are more than one services on a date
        if len(service_list) != 1:
            col1 = get_column_letter(column)
            col2 = get_column_letter(column + len(service_list) - 1)
            ws.merge_cells(f"{col1}1:{col2}1")

        cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)

        # Align to center
        cell.alignment = Alignment(horizontal="center")

        # Set new column based on the number of services
        column += len(service_list)

    return ws

def add_services(ws: Worksheet, services: Services) -> Worksheet:
    """
    Adds service type indicators to the worksheet.

    Args:
        ws (Worksheet): The worksheet to update.
        services (Services): Contains the services for each date.

    Returns:
        Worksheet: The updated worksheet.
    """
    service_list = services_to_list(services)
    # Add services
    for i, item in enumerate(service_list):
        column = get_column_letter(i + 2)
        if item == 0:
            cell = ws.cell(2, i + 2, "o")
            cell.border = Border(top=thick, bottom=thick, left=thick)
            ws.column_dimensions[column].width = 4
        elif item == 1:
            cell = ws.cell(2, i + 2, "a")
            cell.border = Border(top=thick, bottom=thick, right=thick)
            ws.column_dimensions[column].width = 4
        else:
            cell = ws.cell(2, i + 2)
            # Don't overwrite nonempty content
            if cell.value == "":
                cell.value = "tbd"
                cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")

            cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)

            ws.column_dimensions[column].width = 8

        cell.alignment = Alignment(horizontal="center")

    return ws


#############################################
# --- Add counting and preference cells --- #
#############################################

def add_num_times_col(ws: Worksheet, persons: Persons, n_services: int) -> Worksheet:
    """
    Adds a column showing each member's scheduling preference.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.
        n_services (int): Total number of services.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initializations
    n_cols = n_services + 1
    center = Alignment(horizontal="center")

    # Add "Aantal keer" header
    add_header(ws, 1, n_cols + 1, "Aantal keer:", 12)

    # Fill the created columns
    for i in range(len(persons)):
        row = (2 * i) + 3
        # Add counting function for total number of times present
        end_col = get_column_letter(n_cols)
        cell = ws.cell(row, n_cols + 1, f'=COUNTIF(B{row}:{end_col}{row}, "ü")')
        cell.alignment = center

        # Add borders
        if i == 0:
            cell.border = Border(top=thick, left=thick, right=thick)
        else:
            cell.border = Border(top=thin, left=thick, right=thick)

        cell = ws.cell(row + 1, n_cols + 1)
        if i == len(kerkenraad) - 1:
            cell.border = Border(bottom=thick, left=thick, right=thick)
        else:
            cell.border = Border(bottom=thin, left=thick, right=thick)

    return ws

def add_preference_col(ws: Worksheet, persons: Persons, n_services: int) -> Worksheet:
    """
    Adds a column showing each member's scheduling preference.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.
        n_services (int): Total number of services.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initializations
    n_cols = n_services + 1
    center = Alignment(horizontal="center")

    # Add "Voorkeur" header
    add_header(ws, 1, n_cols + 2, "Voorkeur:", 12.5)

    # Fill the created columns
    for i, member in enumerate(persons):
        row = (2 * i) + 3

        # Add preference
        preference = persons[member]["voorkeur"]
        if preference == "geen":
            cell = ws.cell(row, n_cols + 2, '')
        else:
            cell = ws.cell(row, n_cols + 2, f"{preference}")

        # Add styling
        cell.alignment = center

        if i == 0:
            cell.border = Border(top=thick, left=thick, right=thick)
        else:
            cell.border = Border(top=thin, left=thick, right=thick)

        cell = ws.cell(row + 1, n_cols + 2)
        if i == len(persons) - 1:
            cell.border = Border(bottom=thick, left=thick, right=thick)
        else:
            cell.border = Border(bottom=thin, left=thick, right=thick)

    return ws

def add_counting_row(ws: Worksheet, persons: Persons, n_services: int) -> Worksheet:
    """
    Adds a summary row that counts availability indicators per service column.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.
        n_services (int): Total number of services.

    Returns:
        Worksheet: The updated worksheet.
    """
    ending_col = n_services + 2
    last_row = 2*len(persons) + 3

    for i in range(2, ending_col):
        col = get_column_letter(i)
        count_formula = f'=COUNTIF({col}{3}:{col}{last_row-1}, "ü")'
        cell = ws.cell(last_row, i, count_formula)
        cell.alignment = Alignment(horizontal="center")
        if i == ending_col - 1:
            cell.border = Border(right=thick, top=thick, bottom=thick)
        else:
            cell.border = Border(top=thick, bottom=thick)

    return ws


#############################################
# ------------- Add main block ------------ #
#############################################

def add_availability(ws: Worksheet, persons: Persons, services: Services,
                     availability: Optional[np.ndarray] = None) -> Worksheet:
    """
    Adds availability indicators to the worksheet.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.
        services (Services): Contains the services for each date.
        availability (Optional[np.ndarray]): 2D numpy array of availability values.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initializations
    n_services = len(services_to_list(services))

    if availability is None:
        availability = np.zeros((len(persons), n_services), dtype=int)

    # Check for valid input
    assert isinstance(availability, np.ndarray)
    assert len(persons), n_services == availability.shape

    # Loop through all elements
    for i in range(len(persons)):
        # Initialize counter
        j = 0
        for service_list in services.values():
            for service in service_list:
                # Add value to cell
                value = availability[i, j]
                if value == -1:
                    cell = ws.cell(2*i + 3, j + 2, "X")
                elif value == 0:
                    cell = ws.cell(2*i + 3, j + 2, "")
                elif value == 1:
                    cell = ws.cell(2*i + 3, j + 2, "ü")
                    cell.font = Font(name="Wingdings")
                elif value == 2:
                    cell = ws.cell(2*i + 3, j + 2, "~")
                else:
                    pass

                # Add borders
                if service == service_list[0]:
                    cell.border = Border(top=thin, left=thin)
                else:
                    cell.border = Border(top=thin)

                # Align in the center
                cell.alignment = Alignment(horizontal="center")

                # update j
                j += 1

    return ws

def add_tasks(ws: Worksheet, persons: Persons, services: Services,
                     tasks: Optional[np.ndarray] = None) -> Worksheet:
    """
    Adds task descriptions below availability rows.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.
        services (Services): Contains the services for each date.
        tasks (Optional[np.ndarray]): 2D numpy array of task values.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Initializations
    n_services = len(services_to_list(services))

    if tasks is None:
        tasks = np.zeros((len(persons), n_services), dtype=str)

    # Check for valid input
    assert isinstance(tasks, np.ndarray)
    assert len(persons), n_services == tasks.shape

    # Loop through all elements
    for i in range(len(persons)):
        # Initialize counter
        j = 0
        for service_list in services.values():
            for service in service_list:
                # Add task to cell
                task = tasks[i, j]
                cell = ws.cell(2*i + 4, j + 2, task)

                # Add borders
                if service == service_list[0]:
                    cell.border = Border(bottom=thin, left=thin)
                else:
                    cell.border = Border(bottom=thin)

                # Align in the center
                cell.alignment = Alignment(horizontal="center")

                # update j
                j += 1

    return ws


############################################
# ------------ Apply styling ------------- #
############################################

def add_styling_first_col(ws: Worksheet, persons: Persons) -> Worksheet:
    """
    Applies styling to the first column containing names and phone numbers.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.

    Returns:
        Worksheet: The updated worksheet.
    """
    # Apply font and borders for the first column
    for i, row in enumerate(ws.iter_rows(max_col=1)):
        for cell in row:
            cell.font = Font(size=9)
            if i == 2:  # First cell with persons
                cell.border = Border(top=thick, right=thick)
            elif i == 2*len(persons) + 2:  # Last cell
                cell.border = Border(top=thick, right=thick, bottom=thick)
            elif i > 2*len(persons) + 2:  # Anything past the last cell
                pass
            elif i % 2 == 0:
                cell.border = Border(top=thin, right=thick)
            elif i % 2 == 1:
                cell.border = Border(bottom=thin, right=thick)
    return ws

def add_colors(ws: Worksheet, persons: Persons) -> Worksheet:
    """
    Adds background colors for each member's row for better readability.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.

    Returns:
        Worksheet: The updated worksheet.

    Raises:
        Exception: If number of colors does not match the number of persons.
    """
    colors = ["FFD966", "FFF2CC", "FFD966", "FCE4D6", "F4B084", "FCE4D6",
              "9BC2E6", "DDEBF7", "9BC2E6", "E2EFDA", "A9D08E", "E2EFDA", "A9D08E"]

    # Check whether the number of colors is correct
    if len(colors) < len(persons):
        raise ValueError("The number of colors you provided is smaller than the " +
                        "number of members.")
    elif len(colors) > len(persons):
        raise ValueError("The number of colors you provided is greater than the " +
                        "number of members.")

    # Apply colors
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=2*len(persons) + 2)):
        for cell in row:
            color_index = i//2
            cell.fill = PatternFill(fill_type="solid", fgColor=colors[color_index])

    return ws


############################################
# --------------  Add legend ------------- #
############################################

def add_legend(ws: Worksheet, persons: Persons) -> Worksheet:
    """
    Adds a legend section to the bottom of the worksheet.

    Args:
        ws (Worksheet): The worksheet to update.
        persons (Persons): Dictionary of person info.

    Returns:
        Worksheet: The updated worksheet.
    """
    n_rows = 2*len(persons) + 3
    ws.cell(n_rows + 2, 1, "ouderling van dienst =")
    ws.cell(n_rows + 2, 2, "O")

    ws.cell(n_rows + 3, 1, "afkondigingen =")
    ws.cell(n_rows + 3, 2, "A")

    ws.cell(n_rows + 4, 1, "avondmaal Schutse =")
    ws.cell(n_rows + 4, 2, "AS")

    ws.cell(n_rows + 5, 1, "aanwezig =")
    ws.cell(n_rows + 5, 2, "ü")

    ws.cell(n_rows + 6, 1, "gemeld afwezig =")
    ws.cell(n_rows + 6, 2, "X")

    ws.cell(n_rows + 7, 1, "erg voorkeur niet =")
    ws.cell(n_rows + 7, 2, "~")

    ws.cell(n_rows + 8, 1, "beameren =")
    ws.cell(n_rows + 8, 2, "B")

    # Set font size to 9
    for i in range(2, 9):
        cell = ws[f"A{n_rows + i}"]
        cell.font = Font(size=9)
        cell = ws[f"B{n_rows + i}"]
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='center')

    ws[f"B{n_rows + 5}"].font = Font(name="Wingdings", size=9)

    return ws


############################################
# ------------  Main function ------------ #
############################################

def create_excel(file_path: str, persons: Persons, services: Services,
                 availability: np.ndarray = None, tasks: np.ndarray = None) -> None:
    """
    Creates an excel file of the schedule and saves it to disk.

    Args:
        file_path (str): Where to store the excel file.
        persons (Persons): Dictionary of person info.
        services (Services): Contains the services for each date.
        availability (np.ndarray, optional): 2D numpy array of availability
            values. Defaults to None.
        tasks (np.ndarray, optional): 2D numpy array of task values.
            Defaults to None.
    """
    n_services = len(services_to_list(services))

    if os.path.isfile(file_path):
        # Load the workbook
        workbook = openpyxl.load_workbook(filename=file_path)
    else:
        # Create a workbook
        workbook = openpyxl.Workbook()

    # Select the active worksheet
    ws = workbook.active

    # Add column containing persons
    add_first_col(ws, persons)
    autoscale_col(ws, 'A')
    add_styling_first_col(ws, persons)

    # Add dates and services
    add_dates(ws, services)
    add_services(ws, services)

    # Add counting column and preference column
    add_num_times_col(ws, persons, n_services)
    add_preference_col(ws, persons, n_services)

    # Add counting row
    add_counting_row(ws, persons, n_services)

    # Add availability and tasks
    add_availability(ws, persons, services, availability)
    add_tasks(ws, persons, services, tasks)

    # Add colors and legend
    add_colors(ws, persons)
    add_legend(ws, persons)

    # Save workbook
    workbook.save(file_path)
    workbook.close()

    # Open workbook in Excel
    os.system(f"start excel.exe {file_path}")

# Get dates of the schedule
start_date = datetime(2025, 5, 24)
end_date = datetime(2025, 9, 24)
church_dates = get_church_dates(start_date, end_date)
services_dict = get_services(church_dates)

if __name__ == "__main__":
    create_excel("test_kerkenraadsrooster.xlsx", kerkenraad, services_dict)
